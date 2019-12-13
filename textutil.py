import openpyxl
import argparse
import os
from datetime import datetime
from tkinter import *
from tkinter import ttk
from tkinter import filedialog


def print_log(msg):
    print("%s\t%s" % (datetime.now().strftime("%Y-%m-%d %H:%M"), msg))


def process_sap_file(file_name, sheet_name, log_area):
    try:
        print_log("读取%s[%s]" % (file_name, sheet_name))
        xl = openpyxl.load_workbook(file_name)
        sheet = xl[sheet_name]
        print_log("最大行数：%s" % sheet.max_row)
        print_log("开始处理数据")
        hours_by_date_car = {}
        for row in range(sheet.max_row):
            if row == 0:
                continue
            date_info = sheet.cell(row=row+1, column=1).value
            car_no = sheet.cell(row=row+1, column=5).value
            hours = sheet.cell(row=row+1, column=6).value
            if car_no is None or hours is None:
                continue
            date_car = "%s,%s" % (date_info, car_no)
            if date_car in hours_by_date_car:
                hours_by_date_car[date_car] += hours
            else:
                hours_by_date_car[date_car] = hours
        print_log("保存结果")
        new_sheet_name = '%s-修改后' % sheet_name
        while new_sheet_name in xl.sheetnames:
            new_sheet_name = '%s-修改后' % new_sheet_name
        new_sheet = xl.create_sheet(title=new_sheet_name)
        row_index = 1
        for key, value in hours_by_date_car.items():
            date_car = key.split(",")
            new_sheet.cell(row=row_index, column=1).value = date_car[0]
            new_sheet.cell(row=row_index, column=2).value = date_car[1]
            new_sheet.cell(row=row_index, column=3).value = value
            row_index += 1
        xl.save(file_name)
        xl.close()
        log_area.insert("1.0", "处理完毕，结果保存在：[%s]\n" % new_sheet_name)
    except Exception as e:
        print_log("发生异常：%s" % str(e))
        log_area.insert("1.0", "发生异常：%s\n" % str(e))
        return


def create_window():
    root = Tk()
    root.title("文本数据处理小程序")

    note_book = ttk.Notebook(root, padding="1 1 1 1", width=800, height=600)
    note_book.grid(column=0, row=0, sticky=(N, W, E, S))

    sap_frame = ttk.Frame(note_book)
    sap_frame.grid(column=0, row=0, sticky=(N, W, E, S))

    other_frame = ttk.Frame(note_book)
    other_frame.grid(column=0, row=0, sticky=(N, W, E, S))

    root.columnconfigure(0, weight=1)
    root.rowconfigure(0, weight=1)

    note_book.add(sap_frame, text="电源车数据处理")
    note_book.add(other_frame, text="其他数据")

    ttk.Label(sap_frame, text="原始数据文件：").grid(column=1, row=1, sticky=(W, E))
    sap_file_name = StringVar()
    log_area = Text(sap_frame)
    log_area.grid(column=1, row=3, columnspan=3)
    sbar = ttk.Scrollbar(sap_frame, orient=VERTICAL, command=log_area.yview)
    sbar.grid(column=4, row=3, sticky=(N, S))
    log_area.configure(yscrollcommand=sbar.set)

    def select_file():
        nonlocal sap_file_name, log_area
        sap_file_name.set(filedialog.askopenfilename())
        # 太慢了，取消
        # xl = openpyxl.load_workbook(str(sap_file_name.get()))
        # log_area.insert("1.0", "可选的工作簿名称%s\n" % " ".join(["[%s]" % x for x in xl.sheetnames]))
        # xl.close()
    sap_file_entry = ttk.Entry(sap_frame, width=25, textvariable=sap_file_name)
    sap_file_entry.grid(column=2, row=1, sticky=(W, E))

    ttk.Button(sap_frame, text="浏览...", command=select_file).grid(column=3, row=1, sticky=W)

    sheet_name = StringVar()
    ttk.Label(sap_frame, text="工作簿名称：").grid(column=1, row=2, sticky=(W, E))
    sheet_name_entry = ttk.Entry(sap_frame, width=25, textvariable=sheet_name)
    sheet_name_entry.grid(column=2, row=2, sticky=(W, E))

    btn = ttk.Button(sap_frame, text="处理并保存")
    btn.grid(column=3, row=2, sticky=W)
    btn.bind('<1>', lambda e: process_sap_file(str(sap_file_name.get()), str(sheet_name.get()), log_area))

    for child in sap_frame.winfo_children():
        child.grid_configure(padx=5, pady=5)
    sbar.grid_configure(padx=0, pady=0)
    log_area.grid_configure(pady=0, padx=0)

    sap_file_entry.focus()

    root.mainloop()


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='处理文本数据小程序.')
    parser.add_argument('--cmd', nargs='?', help='直接用命令行处理')
    parser.add_argument('--f', nargs='?', help='源文件名')
    parser.add_argument('--s', nargs='?', help='工作簿名称')
    args = parser.parse_args()
    if args.cmd is not None:
        if args.f is None:
            print_log("使用默认文件名：原版.xlsx")
            cmd_file_name = "原版.xlsx"
        else:
            cmd_file_name = args.f
        if not os.path.exists(cmd_file_name):
            print_log("%s不存在。" % cmd_file_name)
            exit()
        if args.s is None:
            month = datetime.now().month
            if month == 1:
                month = 12
            else:
                month = month - 1
            print_log("使用默认的工作簿名称（当前月份减一）：%d" % month)
            cmd_sheet_name = str(month)
        else:
            cmd_sheet_name = args.s
        process_sap_file(cmd_file_name, cmd_sheet_name)
    else:
        create_window()
